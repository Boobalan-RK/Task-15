from openpyxl import load_workbook
from datetime import datetime


class ExcelUtils:
    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path)
        self.sheet = self.wb.active

    def get_data(self):
        rows = []
        for row in range(2, self.sheet.max_row + 1):
            rows.append({
                "row": row,
                "username": self.sheet.cell(row, 2).value,
                "password": self.sheet.cell(row, 3).value
            })
        return rows

    def write_result(self, row, result):
        self.sheet.cell(row, 6).value = result
        self.sheet.cell(row, 4).value = datetime.now().strftime("%Y-%m-%d")
        self.sheet.cell(row, 5).value = "Boobalan"
        self.wb.save(self.path)
